import csv
import json
from tkinter import Tk, filedialog, messagebox, Label, Button, StringVar, OptionMenu
from openpyxl import load_workbook, Workbook # type: ignore


# Function to load file based on format
def load_file(file_path):
    try:
        if file_path.endswith('.csv'):
            with open(file_path, mode='r', newline='') as file:
                reader = csv.reader(file)
                data = [row for row in reader]
        elif file_path.endswith('.xlsx'):
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            data = [[cell.value for cell in row] for row in sheet.iter_rows()]
        elif file_path.endswith('.json'):
            with open(file_path, 'r') as file:
                data = json.load(file)
        else:
            raise ValueError("Unsupported file format!")
        return data
    except FileNotFoundError:
        messagebox.showerror("Error", "File not found. Please check the file path.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while loading the file: {e}")


# Function to preview the first few rows of data
def preview_data(data, num_rows=5):
    try:
        if isinstance(data, list):
            preview = "\n".join([str(row) for row in data[:num_rows]])
        elif isinstance(data, dict):
            preview = json.dumps(data, indent=4)[:num_rows * 100]
        else:
            preview = "Unsupported data format for preview."
        messagebox.showinfo("Preview", preview)
    except Exception as e:
        messagebox.showerror("Error", f"Error while previewing data: {e}")


# Function to convert data to a new format and save
def convert_file(data, output_format, output_path):
    try:
        if output_format == 'csv':
            with open(output_path, mode='w', newline='') as file:
                writer = csv.writer(file)
                if isinstance(data, list):
                    writer.writerows(data)
                else:
                    raise ValueError("Data format not suitable for CSV conversion.")
        elif output_format == 'excel':
            workbook = Workbook()
            sheet = workbook.active
            if isinstance(data, list):
                for row in data:
                    sheet.append(row)
            workbook.save(output_path)
        elif output_format == 'json':
            with open(output_path, 'w') as file:
                json.dump(data, file, indent=4)
        else:
            raise ValueError("Unsupported output format!")
        messagebox.showinfo("Success", f"File saved as {output_format} at {output_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Error during conversion: {e}")


# GUI Setup
def file_conversion_tool_gui():
    def load_input_file():
        file_path = filedialog.askopenfilename(filetypes=[("All Files", "."),
                                                          ("CSV Files", "*.csv"),
                                                          ("Excel Files", "*.xlsx"),
                                                          ("JSON Files", "*.json")])
        if file_path:
            input_file_var.set(file_path)
            global data
            data = load_file(file_path)
            if data:
                preview_button.config(state="normal")

    def save_output_file():
        file_path = filedialog.asksaveasfilename(defaultextension=f".{output_format_var.get()}",
                                                 filetypes=[("CSV Files", "*.csv"),
                                                            ("Excel Files", "*.xlsx"),
                                                            ("JSON Files", "*.json")])
        if file_path:
            output_file_var.set(file_path)

    def convert_file_gui():
        if not data:
            messagebox.showerror("Error", "No data loaded. Please load a file first.")
            return
        output_format = output_format_var.get()
        output_path = output_file_var.get()
        if not output_path:
            messagebox.showerror("Error", "Please specify an output file path.")
            return
        convert_file(data, output_format, output_path)

    root = Tk()
    root.title("File Conversion Tool")

    global input_file_var, output_file_var, output_format_var, data
    input_file_var = StringVar()
    output_file_var = StringVar()
    output_format_var = StringVar(value="csv")
    data = None

    # Input File
    Label(root, text="Input File:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
    Label(root, textvariable=input_file_var, relief="solid", width=40).grid(row=0, column=1, padx=10, pady=10)
    Button(root, text="Browse", command=load_input_file).grid(row=0, column=2, padx=10, pady=10)

    # Preview Button
    preview_button = Button(root, text="Preview", command=lambda: preview_data(data), state="disabled")
    preview_button.grid(row=1, column=1, padx=10, pady=10)

    # Output Format
    Label(root, text="Output Format:").grid(row=2, column=0, padx=10, pady=10, sticky="w")
    OptionMenu(root, output_format_var, "csv", "excel", "json").grid(row=2, column=1, padx=10, pady=10, sticky="w")

    # Output File
    Label(root, text="Output File:").grid(row=3, column=0, padx=10, pady=10, sticky="w")
    Label(root, textvariable=output_file_var, relief="solid", width=40).grid(row=3, column=1, padx=10, pady=10)
    Button(root, text="Save As", command=save_output_file).grid(row=3, column=2, padx=10, pady=10)

    # Convert Button
    Button(root, text="Convert", command=convert_file_gui).grid(row=4, column=1, pady=20)

    root.mainloop()


# Running the GUI
if _name_ == "_main_":
    file_conversion_tool_gui()